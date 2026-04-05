import io
import base64
from datetime import datetime
from backend.logger import get_logger

logger = get_logger(__name__)

# ── ReportLab imports at module level (required by helper functions) ──────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm, mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, HRFlowable, KeepTogether, PageBreak
    )
    from reportlab.graphics.shapes import Drawing, Rect, String, Circle, Line
    from reportlab.graphics import renderPDF
except ImportError as e:
    raise RuntimeError(f"reportlab not installed: {e}")

# ── Brand Colors ──────────────────────────────────────────────────────────────
BRAND_DARK   = "#0f1117"
BRAND_NAVY   = "#1a1f2e"
BRAND_BLUE   = "#4f9cff"
BRAND_CYAN   = "#22d3ee"
BRAND_GREEN  = "#10b981"
BRAND_AMBER  = "#f59e0b"
BRAND_ROSE   = "#f43f5e"
BRAND_VIOLET = "#8b5cf6"
BRAND_LIGHT  = "#f8fafc"
BRAND_MUTED  = "#94a3b8"
TEXT_DARK    = "#1e293b"
TEXT_MID     = "#475569"

W, H = A4


def _hex(h):
    return colors.HexColor(h)


def generate_pdf_report(result: dict, chart_images: list) -> bytes:
    """Generate a professional, visually rich PDF report."""
    buffer = io.BytesIO()

    # ── Page template with header/footer ─────────────────────────────────────
    def _on_page(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(_hex(BRAND_BLUE))
        canvas.rect(0, H - 6*mm, W, 6*mm, fill=1, stroke=0)
        canvas.setFillColor(_hex(BRAND_NAVY))
        canvas.rect(0, 0, W, 10*mm, fill=1, stroke=0)
        canvas.setFillColor(_hex(BRAND_MUTED))
        canvas.setFont("Helvetica", 7)
        canvas.drawString(2*cm, 3.5*mm, "SparkInsight Analytics Platform  •  Confidential")
        canvas.drawRightString(W - 2*cm, 3.5*mm, f"Page {doc.page}  •  {datetime.now().strftime('%Y-%m-%d')}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2.2*cm, bottomMargin=1.8*cm,
        title="SparkInsight Analytics Report",
        author="SparkInsight Platform",
    )

    # ── Styles ────────────────────────────────────────────────────────────────
    S = getSampleStyleSheet()

    def style(name, **kw):
        return ParagraphStyle(name, parent=S["Normal"], **kw)

    sTitle    = style("sTitle",    fontSize=26, textColor=_hex(BRAND_LIGHT),    fontName="Helvetica-Bold",  leading=32, spaceAfter=4)
    sSubtitle = style("sSub",      fontSize=13, textColor=_hex(BRAND_CYAN),     fontName="Helvetica",       leading=18, spaceAfter=2)
    sMeta     = style("sMeta",     fontSize=9,  textColor=_hex(BRAND_MUTED),    fontName="Helvetica",       leading=13)
    sH2       = style("sH2",       fontSize=13, textColor=_hex(BRAND_BLUE),     fontName="Helvetica-Bold",  leading=18, spaceBefore=10, spaceAfter=6)
    sH3       = style("sH3",       fontSize=10, textColor=_hex(TEXT_DARK),      fontName="Helvetica-Bold",  leading=14, spaceAfter=4)
    sBody     = style("sBody",     fontSize=9,  textColor=_hex(TEXT_MID),       fontName="Helvetica",       leading=13, spaceAfter=3)
    sInsight  = style("sInsight",  fontSize=9,  textColor=_hex(TEXT_DARK),      fontName="Helvetica",       leading=13)
    sLabel    = style("sLabel",    fontSize=8,  textColor=_hex(BRAND_MUTED),    fontName="Helvetica",       leading=11)
    sCentered = style("sCentered", fontSize=9,  textColor=_hex(TEXT_MID),       fontName="Helvetica",       leading=13, alignment=TA_CENTER)

    story = []

    # ═══════════════════════════════════════════════════════════════════════════
    # COVER PAGE
    # ═══════════════════════════════════════════════════════════════════════════
    ds  = result.get("dataset_info", {})
    job = result.get("job_info", {})
    quality = result.get("quality_score", {})

    cover_bg = Table(
        [[Paragraph("⚡ SparkInsight Analytics Platform", sTitle)],
         [Paragraph(f"Analytics Report", sSubtitle)],
         [Spacer(1, 4*mm)],
         [Paragraph(f"Dataset: {(ds.get('dataset_name','Unknown')).replace('_',' ').title()}", sSubtitle)],
         [Spacer(1, 2*mm)],
         [Paragraph(f"Category: {ds.get('category','').title()}  •  Format: {ds.get('format','').upper()}  •  Size: {ds.get('estimated_size','')}  •  Version: {ds.get('version','')}", sMeta)],
         [Spacer(1, 2*mm)],
         [Paragraph(f"Job ID: {job.get('job_id','')}  •  Generated: {datetime.now().strftime('%B %d, %Y at %H:%M UTC')}", sMeta)],
        ],
        colWidths=[W - 4*cm],
    )
    cover_bg.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,-1), _hex(BRAND_NAVY)),
        ("LEFTPADDING",  (0,0), (-1,-1), 16),
        ("RIGHTPADDING", (0,0), (-1,-1), 16),
        ("TOPPADDING",   (0,0), (-1,-1), 16),
        ("BOTTOMPADDING",(0,0), (-1,-1), 8),
        ("ROUNDEDCORNERS", (0,0), (-1,-1), [8,8,8,8]),
    ]))
    story.append(cover_bg)
    story.append(Spacer(1, 6*mm))

    # ── Cover KPI Row ─────────────────────────────────────────────────────────
    sm = result.get("summary_metrics", {})
    em = result.get("execution_metrics", {})
    cs = result.get("cleaning_summary", {})

    def kpi_cell(label, value, color=BRAND_BLUE):
        return [
            Paragraph(f'<font color="{color}"><b>{value}</b></font>', style("kpiv", fontSize=18, fontName="Helvetica-Bold", leading=22, alignment=TA_CENTER, textColor=_hex(color))),
            Paragraph(label, style("kpil", fontSize=8, fontName="Helvetica", leading=11, alignment=TA_CENTER, textColor=_hex(BRAND_MUTED))),
        ]

    score = quality.get("overall_score", 0)
    score_color = BRAND_GREEN if score >= 80 else BRAND_AMBER if score >= 60 else BRAND_ROSE

    kpi_table = Table([
        [
            Table([kpi_cell("Total Records",     _fmt_num(sm.get("total_records", 0)),  BRAND_BLUE)],   colWidths=[3.8*cm]),
            Table([kpi_cell("Quality Score",     f"{score}/100",                         score_color)],  colWidths=[3.8*cm]),
            Table([kpi_cell("Exec Time",         _fmt_secs(em.get("total_execution_time_sec")), BRAND_CYAN)], colWidths=[3.8*cm]),
            Table([kpi_cell("Rows After Clean",  _fmt_num(cs.get("final_rows", 0)),      BRAND_VIOLET)], colWidths=[3.8*cm]),
        ]
    ], colWidths=[3.8*cm]*4)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), _hex(BRAND_NAVY)),
        ("ROUNDEDCORNERS",(0,0),(-1,-1),[6,6,6,6]),
        ("TOPPADDING",   (0,0), (-1,-1), 10),
        ("BOTTOMPADDING",(0,0), (-1,-1), 10),
        ("LEFTPADDING",  (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("LINEBEFORE",   (1,0), (3,0), 0.5, _hex("#2d3748")),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 6*mm))

    desc = ds.get("description", "")
    if desc:
        story.append(Paragraph(desc, sBody))
        story.append(Spacer(1, 4*mm))

    story.append(PageBreak())

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 1 — DATA QUALITY ASSESSMENT
    # ═══════════════════════════════════════════════════════════════════════════
    story.append(_section_header("01", "Data Quality Assessment", BRAND_GREEN))
    story.append(Spacer(1, 3*mm))

    comp = quality.get("component_breakdown", {})
    weights = quality.get("weights_used", {})
    grade = quality.get("grade", "?")

    score_badge = Table([
        [Paragraph(f'<font color="{score_color}"><b>{score}</b></font>',
                   style("qs", fontSize=40, fontName="Helvetica-Bold", leading=44, alignment=TA_CENTER, textColor=_hex(score_color)))],
        [Paragraph("/ 100", style("qs2", fontSize=11, fontName="Helvetica", leading=14, alignment=TA_CENTER, textColor=_hex(BRAND_MUTED)))],
        [Paragraph(f"Grade  {grade}", style("qs3", fontSize=12, fontName="Helvetica-Bold", leading=16, alignment=TA_CENTER, textColor=_hex(score_color)))],
    ], colWidths=[5*cm])
    score_badge.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), _hex(BRAND_NAVY)),
        ("TOPPADDING",   (0,0), (-1,-1), 16),
        ("BOTTOMPADDING",(0,0), (-1,-1), 16),
        ("ROUNDEDCORNERS",(0,0),(-1,-1),[8,8,8,8]),
    ]))

    comp_rows = [["Component", "Score", "Weight"]]
    for k, v in comp.items():
        comp_rows.append([
            k.replace("_", " ").title(),
            f"{v}",
            f"{int(weights.get(k,0)*100)}%",
        ])

    comp_table = _styled_table(comp_rows, col_widths=[7*cm, 3*cm, 2.5*cm], highlight_last=False)

    quality_row = Table([[score_badge, comp_table]], colWidths=[5.5*cm, 11*cm])
    quality_row.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"), ("LEFTPADDING",(1,0),(1,0),12)]))
    story.append(quality_row)
    story.append(Spacer(1, 4*mm))

    recs = quality.get("quality_recommendations", [])
    if recs:
        story.append(Paragraph("Recommendations", sH3))
        for r in recs:
            story.append(Paragraph(f"<font color='{BRAND_AMBER}'>▶</font>  {r}", sBody))
    story.append(Spacer(1, 6*mm))

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 2 — PROFILING & CLEANING
    # ═══════════════════════════════════════════════════════════════════════════
    story.append(_section_header("02", "Data Profiling & Cleaning", BRAND_CYAN))
    story.append(Spacer(1, 3*mm))

    profile = result.get("profiling_report", {})

    prof_data = [
        ["Metric", "Value", "Metric", "Value"],
        ["Total Rows",         _fmt_int(profile.get("row_count",0)),             "Total Columns",     str(profile.get("column_count",0))],
        ["Numeric Columns",    str(len(profile.get("numeric_columns",[]))),      "Categorical Cols",  str(len(profile.get("categorical_columns",[])))],
        ["Date Columns",       str(len(profile.get("date_columns",[]))),         "High Cardinality",  str(len(profile.get("high_cardinality_columns",[])))],
        ["Rows Before Clean",  _fmt_int(cs.get("initial_rows",0)),               "Rows After Clean",  _fmt_int(cs.get("final_rows",0))],
        ["Duplicates Removed", _fmt_int(cs.get("rows_removed",0)),               "Columns Dropped",   str(cs.get("columns_removed",0))],
    ]
    story.append(_styled_table(prof_data, col_widths=[5*cm, 4*cm, 5*cm, 4*cm], dual_header=True))
    story.append(Spacer(1, 6*mm))

    col_profiles = profile.get("column_profiles", [])
    if col_profiles:
        story.append(Paragraph("Column Profiles", sH3))
        cp_rows = [["Column", "Type", "Null %", "Unique", "Min", "Max", "Mean", "Skew"]]
        for c in col_profiles[:15]:
            ns = c.get("numeric_stats") or {}
            sk = c.get("skew_info") or {}
            cp_rows.append([
                c["column"][:22],
                c["column_type"],
                f"{c.get('null_pct',0)}%",
                _fmt_int(c.get("unique_count",0)),
                _fmt_num(ns.get("min")) if ns.get("min") is not None else "—",
                _fmt_num(ns.get("max")) if ns.get("max") is not None else "—",
                _fmt_num(ns.get("mean")) if ns.get("mean") is not None else "—",
                sk.get("direction","sym")[:8] if sk.get("is_skewed") else "sym",
            ])
        story.append(_styled_table(cp_rows, col_widths=[3.8*cm,2*cm,1.5*cm,1.5*cm,2*cm,2*cm,2*cm,1.7*cm]))
        story.append(Spacer(1, 6*mm))

    story.append(PageBreak())

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 3 — SUMMARY METRICS
    # ═══════════════════════════════════════════════════════════════════════════
    story.append(_section_header("03", "Summary Metrics", BRAND_VIOLET))
    story.append(Spacer(1, 3*mm))

    metric_name = sm.get("primary_metric", "Primary Metric")
    sm_rows = [
        ["Metric", "Value"],
        ["Primary Metric Column", metric_name],
        ["Total Records",         _fmt_int(sm.get("total_records", 0))],
        ["Sum",                   _fmt_num(sm.get("total"))],
        ["Average",               _fmt_num(sm.get("average"))],
        ["Minimum",               _fmt_num(sm.get("minimum"))],
        ["Maximum",               _fmt_num(sm.get("maximum"))],
        ["Std Deviation",         _fmt_num(sm.get("std_deviation"))],
    ]
    story.append(_styled_table(sm_rows, col_widths=[8*cm, 8.5*cm]))
    story.append(Spacer(1, 4*mm))

    top10 = sm.get("top_10_records", [])
    if top10 and isinstance(top10, list) and len(top10) > 0:
        story.append(Paragraph(f"Top 10 by {metric_name}", sH3))
        keys = list(top10[0].keys())[:6]
        t10_rows = [keys] + [[str(r.get(k,""))[:20] for k in keys] for r in top10[:10]]
        story.append(_styled_table(t10_rows))
    story.append(Spacer(1, 6*mm))

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 4 — ANALYTICS CHARTS
    # ═══════════════════════════════════════════════════════════════════════════
    if chart_images:
        story.append(_section_header("04", "Analytics Charts", BRAND_BLUE))
        story.append(Spacer(1, 3*mm))

        paired = []
        for i, img_data in enumerate(chart_images):
            title = img_data.get("title", f"Chart {i+1}").replace("-", " ").replace("_", " ").title()
            b64   = img_data.get("image", "")
            if not b64:
                continue
            try:
                img_bytes = base64.b64decode(b64.split(",")[-1])
                img_io    = io.BytesIO(img_bytes)
                chart_w   = 7.8*cm
                chart_h   = 5.5*cm
                img_obj   = Image(img_io, width=chart_w, height=chart_h)

                cell = Table([
                    [Paragraph(title, style("ctitle", fontSize=8, fontName="Helvetica-Bold", leading=11, textColor=_hex(BRAND_BLUE)))],
                    [img_obj],
                ], colWidths=[chart_w])
                cell.setStyle(TableStyle([
                    ("BACKGROUND",   (0,0), (-1,-1), _hex(BRAND_NAVY)),
                    ("TOPPADDING",   (0,0), (-1,-1), 6),
                    ("BOTTOMPADDING",(0,0), (-1,-1), 6),
                    ("LEFTPADDING",  (0,0), (-1,-1), 6),
                    ("RIGHTPADDING", (0,0), (-1,-1), 6),
                    ("ROUNDEDCORNERS",(0,0),(-1,-1),[6,6,6,6]),
                ]))
                paired.append(cell)
            except Exception as ex:
                logger.warning(f"Could not embed chart '{title}': {ex}")

        for i in range(0, len(paired), 2):
            row_cells = paired[i:i+2]
            if len(row_cells) == 1:
                row_cells.append(Spacer(7.8*cm, 1))
            row = Table([row_cells], colWidths=[8.3*cm, 8.3*cm])
            row.setStyle(TableStyle([("LEFTPADDING",(1,0),(1,0),10)]))
            story.append(row)
            story.append(Spacer(1, 4*mm))

        story.append(Spacer(1, 4*mm))

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 5 — CATEGORY ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    cat = result.get("category_analysis", {})
    cats = cat.get("categories", [])
    if cats:
        story.append(PageBreak())
        story.append(_section_header("05", f"Category Analysis — {cat.get('column','')}", BRAND_AMBER))
        story.append(Spacer(1, 3*mm))

        has_total = "total" in cats[0]
        cat_rows = [["Rank", "Category", "Count", "Percentage"] + (["Total", "Average"] if has_total else [])]
        for c in cats[:20]:
            cat_rows.append([
                f"#{c.get('rank', '')}",
                str(c.get("category",""))[:30],
                _fmt_int(c.get("count",0)),
                f"{c.get('percentage',0)}%",
            ] + ([_fmt_num(c.get("total")), _fmt_num(c.get("average"))] if has_total else []))
        story.append(_styled_table(cat_rows))
        story.append(Spacer(1, 6*mm))

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 6 — TIME ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    time_a = result.get("time_analysis", {})
    if time_a.get("available"):
        story.append(_section_header("06", "Time Series Analysis", BRAND_CYAN))
        story.append(Spacer(1, 3*mm))

        monthly = time_a.get("monthly", [])
        yearly  = time_a.get("yearly", [])

        if monthly:
            story.append(Paragraph("Monthly Aggregation", sH3))
            m_rows = [["Period", "Total", "Count"]] + [
                [r.get("period",""), _fmt_num(r.get("total")), _fmt_int(r.get("count",0))]
                for r in monthly[:24]
            ]
            story.append(_styled_table(m_rows, col_widths=[6*cm, 6*cm, 5*cm]))
            story.append(Spacer(1, 4*mm))

        if yearly:
            story.append(Paragraph("Yearly Summary", sH3))
            y_rows = [["Year", "Total", "Count", "Growth Rate"]] + [
                [r.get("year",""), _fmt_num(r.get("total")), _fmt_int(r.get("count",0)),
                 f"{r.get('growth_rate','')}%" if r.get("growth_rate") is not None else "—"]
                for r in yearly
            ]
            story.append(_styled_table(y_rows, col_widths=[4*cm, 5*cm, 4*cm, 5*cm]))
        story.append(Spacer(1, 6*mm))

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 7 — CORRELATION ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    corr = result.get("correlation_analysis", {})
    strong = corr.get("strong_correlations", [])
    if strong:
        story.append(_section_header("07", "Correlation Analysis", BRAND_ROSE))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph("Strong Correlations Detected", sH3))
        sc_rows = [["Column A", "Column B", "Correlation", "Strength"]]
        for s in strong[:10]:
            val = s.get("correlation", 0)
            sc_rows.append([
                s.get("col1",""), s.get("col2",""),
                f"{val:.3f}",
                s.get("strength",""),
            ])
        story.append(_styled_table(sc_rows, col_widths=[5*cm, 5*cm, 3.5*cm, 4*cm]))
        story.append(Spacer(1, 6*mm))

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 8 — EXECUTION METRICS
    # ═══════════════════════════════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(_section_header("08", "Execution Metrics", BRAND_VIOLET))
    story.append(Spacer(1, 3*mm))

    em_rows = [
        ["Metric", "Value", "Metric", "Value"],
        ["Total Execution Time",   _fmt_secs(em.get("total_execution_time_sec")),    "Records Processed",  _fmt_int(em.get("records_processed",0))],
        ["Records / Second",       _fmt_int(em.get("records_per_second",0)),          "Partition Count",    str(em.get("partition_count","—"))],
        ["Spark Version",          str(em.get("spark_version","—")),                  "Adaptive Execution", "✓ Enabled" if em.get("adaptive_execution_enabled") else "✗ Disabled"],
        ["Dataset Size",           _fmt_bytes(em.get("dataset_size_bytes")),           "Shuffle Read",       _fmt_bytes(em.get("shuffle_read_bytes"))],
    ]
    story.append(_styled_table(em_rows, col_widths=[5*cm, 4*cm, 5*cm, 4*cm], dual_header=True))
    story.append(Spacer(1, 4*mm))

    stages = em.get("stage_times", {})
    if stages:
        story.append(Paragraph("Pipeline Stage Breakdown", sH3))
        st_rows = [["Stage", "Time (s)"]] + [[k.replace("_"," ").title(), str(v)] for k,v in stages.items()]
        story.append(_styled_table(st_rows, col_widths=[12*cm, 5.5*cm]))
    story.append(Spacer(1, 6*mm))

    # ═══════════════════════════════════════════════════════════════════════════
    # SECTION 9 — INSIGHTS
    # ═══════════════════════════════════════════════════════════════════════════
    insights = result.get("insights", [])
    if insights:
        story.append(_section_header("09", "Automated Insights", BRAND_GREEN))
        story.append(Spacer(1, 3*mm))

        sev_icon  = {"success": "✅", "warning": "⚠️", "info": "ℹ️", "error": "❌"}
        sev_color = {"success": BRAND_GREEN, "warning": BRAND_AMBER, "info": BRAND_BLUE, "error": BRAND_ROSE}

        for ins in insights:
            sev   = ins.get("severity", "info")
            icon  = sev_icon.get(sev, "ℹ️")
            color = sev_color.get(sev, BRAND_BLUE)
            ins_cell = Table([
                [Paragraph(f'<font color="{color}"><b>{icon}  {ins.get("title","")}</b></font>', sInsight)],
                [Paragraph(ins.get("detail",""), sBody)],
            ], colWidths=[W - 4*cm - 6*mm])
            ins_cell.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,-1), _hex(BRAND_NAVY)),
                ("LEFTPADDING",  (0,0), (-1,-1), 10),
                ("RIGHTPADDING", (0,0), (-1,-1), 10),
                ("TOPPADDING",   (0,0), (-1,-1), 7),
                ("BOTTOMPADDING",(0,0), (-1,-1), 7),
                ("LINEAFTER",    (0,0), (0,-1), 3, _hex(color)),
            ]))
            story.append(ins_cell)
            story.append(Spacer(1, 3*mm))

    # ── Build ──────────────────────────────────────────────────────────────────
    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    pdf_bytes = buffer.getvalue()
    logger.info(f"PDF report generated: {len(pdf_bytes):,} bytes")
    return pdf_bytes


# ── Helpers ────────────────────────────────────────────────────────────────────

def _section_header(num, title, color=BRAND_BLUE):
    """Colored section header bar — uses module-level cm import."""
    cell = Table([[
        _para(f'<font color="{color}"><b>{num}</b></font>', 16, "Helvetica-Bold", color),
        _para(f'  <b>{title}</b>', 13, "Helvetica-Bold", BRAND_LIGHT),
    ]], colWidths=[1.2*cm, 15*cm])          # ← cm is now available at module level
    cell.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), _hex(BRAND_NAVY)),
        ("TOPPADDING",   (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0), (-1,-1), 8),
        ("LEFTPADDING",  (0,0), (-1,-1), 12),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("LINEBEFORE",   (0,0), (0,-1), 4, _hex(color)),
    ]))
    return cell


def _para(text, size=9, font="Helvetica", color=TEXT_MID, align="LEFT"):
    alignmap = {"LEFT": TA_LEFT, "CENTER": TA_CENTER, "RIGHT": TA_RIGHT}
    s = ParagraphStyle("p", fontName=font, fontSize=size, textColor=_hex(color),
                       leading=size+4, alignment=alignmap.get(align, TA_LEFT))
    return Paragraph(text, s)


def _styled_table(data, col_widths=None, highlight_last=False, dual_header=False):
    if not data:
        return Paragraph("No data available.", _para("", 9).__class__)

    n_cols = len(data[0])
    if col_widths is None:
        col_widths = [(16.6 / n_cols) * cm] * n_cols    # ← cm available here too

    ts = [
        ("BACKGROUND",   (0,0), (-1,0), _hex(BRAND_NAVY)),
        ("TEXTCOLOR",    (0,0), (-1,0), _hex(BRAND_LIGHT)),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,-1), 8.5),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [_hex("#ffffff"), _hex("#f1f5f9")]),
        ("TEXTCOLOR",    (0,1), (-1,-1), _hex(TEXT_DARK)),
        ("FONTNAME",     (0,1), (-1,-1), "Helvetica"),
        ("GRID",         (0,0), (-1,-1), 0.3, _hex("#e2e8f0")),
        ("LINEBELOW",    (0,0), (-1,0), 1.5, _hex(BRAND_BLUE)),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 7),
        ("RIGHTPADDING", (0,0), (-1,-1), 7),
        ("TOPPADDING",   (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",(0,0), (-1,-1), 5),
    ]
    if dual_header:
        ts.append(("BACKGROUND", (2,0), (2,-1), _hex(BRAND_NAVY)))
        ts.append(("FONTNAME",   (2,0), (2,-1), "Helvetica-Bold"))
        ts.append(("TEXTCOLOR",  (2,0), (2,-1), _hex(BRAND_LIGHT)))
    if highlight_last:
        ts.append(("BACKGROUND", (0,-1), (-1,-1), _hex("#ecfdf5")))
        ts.append(("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold"))
        ts.append(("TEXTCOLOR",  (0,-1), (-1,-1), _hex(BRAND_GREEN)))

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(ts))
    return t


def _fmt_num(n):
    if n is None or n == "":
        return "—"
    try:
        v = float(n)
        if abs(v) >= 1e9: return f"{v/1e9:.2f}B"
        if abs(v) >= 1e6: return f"{v/1e6:.2f}M"
        if abs(v) >= 1e3: return f"{v/1e3:.1f}K"
        return f"{v:,.2f}"
    except Exception:
        return str(n)


def _fmt_int(n):
    if n is None: return "—"
    try: return f"{int(n):,}"
    except Exception: return str(n)


def _fmt_secs(s):
    if s is None: return "—"
    try:
        s = float(s)
        if s < 60: return f"{s:.1f}s"
        return f"{int(s//60)}m {int(s%60)}s"
    except Exception: return str(s)


def _fmt_bytes(b):
    if not b: return "—"
    try:
        b = float(b)
        if b >= 1e9: return f"{b/1e9:.2f} GB"
        if b >= 1e6: return f"{b/1e6:.1f} MB"
        return f"{b/1024:.0f} KB"
    except Exception: return str(b)


# ── Excel Report ──────────────────────────────────────────────────────────────

def generate_excel_report(result: dict) -> bytes:
    """Generate multi-sheet Excel report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(f"openpyxl not installed: {e}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL  = PatternFill("solid", fgColor="1a1f2e")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    ALT_FILL     = PatternFill("solid", fgColor="f1f5f9")
    TITLE_FONT   = Font(size=15, bold=True, color="1a1f2e", name="Calibri")
    BORDER       = Border(
        left=Side(style="thin", color="e2e8f0"),
        right=Side(style="thin", color="e2e8f0"),
        top=Side(style="thin", color="e2e8f0"),
        bottom=Side(style="thin", color="e2e8f0"),
    )

    def add_sheet(name, rows, title=None, freeze=True):
        ws = wb.create_sheet(title=name[:31])
        r_offset = 0
        if title:
            ws.cell(1, 1, title).font = TITLE_FONT
            ws.cell(1, 1).fill = PatternFill("solid", fgColor="f8fafc")
            ws.row_dimensions[1].height = 24
            ws.append([])
            r_offset = 2
        for ri, row in enumerate(rows):
            ws.append(row)
            actual_row = ri + 1 + r_offset
            for ci, val in enumerate(row, 1):
                cell = ws.cell(actual_row, ci)
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if ri == 0:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif ri % 2 == 0:
                    cell.fill = ALT_FILL
            ws.row_dimensions[actual_row].height = 16

        for ci in range(1, (len(rows[0]) if rows else 1) + 1):
            col_letter = get_column_letter(ci)
            max_len = max((len(str(r[ci-1])) if ci <= len(r) else 0 for r in rows), default=10)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

        if freeze and rows:
            ws.freeze_panes = ws.cell(2 + r_offset, 1)
        ws.sheet_view.showGridLines = True
        return ws

    ds      = result.get("dataset_info", {})
    profile = result.get("profiling_report", {})
    quality = result.get("quality_score", {})
    cs      = result.get("cleaning_summary", {})
    sm      = result.get("summary_metrics", {})
    cat     = result.get("category_analysis", {})
    time_a  = result.get("time_analysis", {})
    dist    = result.get("distribution_analysis", {})
    corr    = result.get("correlation_analysis", {})
    em      = result.get("execution_metrics", {})
    insights= result.get("insights", [])

    add_sheet("Overview", [
        ["Field", "Value"],
        ["Dataset Name",      ds.get("dataset_name","")],
        ["Category",          ds.get("category","")],
        ["Version",           ds.get("version","")],
        ["Format",            ds.get("format","")],
        ["Estimated Size",    ds.get("estimated_size","")],
        ["Primary Metric",    ds.get("primary_metric","")],
        ["Time Column",       ds.get("time_column","")],
        ["Description",       ds.get("description","")],
        ["Total Rows",        profile.get("row_count",0)],
        ["Total Columns",     profile.get("column_count",0)],
        ["Quality Score",     f"{quality.get('overall_score',0)}/100  (Grade {quality.get('grade','?')})"],
        ["Total Records",     sm.get("total_records",0)],
        ["Exec Time (s)",     em.get("total_execution_time_sec","")],
        ["Records/sec",       em.get("records_per_second","")],
        ["Generated At",      datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ], title="SparkInsight Analytics Report — Overview")

    comp = quality.get("component_breakdown", {})
    w    = quality.get("weights_used", {})
    add_sheet("Quality Score", [
        ["Component", "Score", "Weight", "Status"],
        *[[k.replace("_"," ").title(), v, f"{int(w.get(k,0)*100)}%",
           "✅ Good" if v>=80 else "⚠️ Fair" if v>=60 else "❌ Poor"] for k,v in comp.items()],
        ["OVERALL SCORE", quality.get("overall_score",""), "", f"Grade {quality.get('grade','?')}"],
    ], title="Data Quality Assessment")

    cp = profile.get("column_profiles", [])
    add_sheet("Column Profiles", [
        ["Column", "Type", "Null %", "Unique Count", "Min", "Max", "Mean", "StdDev", "Skewed?"],
        *[[c["column"], c["column_type"], c.get("null_pct",""), c.get("unique_count",""),
           (c.get("numeric_stats") or {}).get("min",""),
           (c.get("numeric_stats") or {}).get("max",""),
           (c.get("numeric_stats") or {}).get("mean",""),
           (c.get("numeric_stats") or {}).get("stddev",""),
           "Yes" if (c.get("skew_info") or {}).get("is_skewed") else "No"] for c in cp],
    ], title="Data Profiling — Column Profiles")

    add_sheet("Summary Metrics", [
        ["Metric", "Value"],
        ["Primary Metric", sm.get("primary_metric","")],
        ["Total Records",  sm.get("total_records","")],
        ["Sum",            sm.get("total","")],
        ["Average",        sm.get("average","")],
        ["Minimum",        sm.get("minimum","")],
        ["Maximum",        sm.get("maximum","")],
        ["Std Deviation",  sm.get("std_deviation","")],
    ], title="Summary Metrics")

    cats = cat.get("categories", [])
    if cats:
        has_total = "total" in cats[0]
        add_sheet("Category Analysis", [
            ["Rank", "Category", "Count", "Percentage"] + (["Total", "Average"] if has_total else []),
            *[[c.get("rank",""), c.get("category",""), c.get("count",""), f"{c.get('percentage','')}%"]
              + ([c.get("total",""), c.get("average","")] if has_total else []) for c in cats],
        ], title=f"Category Analysis — {cat.get('column','')}")

    if time_a.get("available"):
        monthly = time_a.get("monthly", [])
        if monthly:
            add_sheet("Time Analysis", [
                ["Period", "Total", "Count"],
                *[[r.get("period",""), r.get("total",""), r.get("count","")] for r in monthly],
            ], title="Monthly Time Series")

    if dist.get("available") and dist.get("histogram_bins"):
        add_sheet("Distribution", [
            ["Range Start", "Range End", "Count", "Percentage"],
            *[[b.get("range_start",""), b.get("range_end",""), b.get("count",""), b.get("percentage","")] for b in dist["histogram_bins"]],
        ], title=f"Distribution — {dist.get('column','')}")

    strong_corr = corr.get("strong_correlations", [])
    if strong_corr:
        add_sheet("Correlations", [
            ["Column A", "Column B", "Correlation", "Strength"],
            *[[c.get("col1",""), c.get("col2",""), c.get("correlation",""), c.get("strength","")] for c in strong_corr],
        ], title="Correlation Analysis")

    stages = em.get("stage_times", {})
    em_rows = [
        ["Metric", "Value"],
        ["Total Execution Time (s)", em.get("total_execution_time_sec","")],
        ["Records Processed",        em.get("records_processed","")],
        ["Records / Second",         em.get("records_per_second","")],
        ["Partition Count",          em.get("partition_count","")],
        ["Spark Version",            em.get("spark_version","")],
        ["Adaptive Execution",       em.get("adaptive_execution_enabled","")],
        ["Dataset Size (bytes)",     em.get("dataset_size_bytes","")],
        ["Shuffle Read (bytes)",     em.get("shuffle_read_bytes","")],
        ["Shuffle Write (bytes)",    em.get("shuffle_write_bytes","")],
    ] + [["Stage: " + k, v] for k,v in stages.items()]
    add_sheet("Execution Metrics", em_rows, title="Spark Execution Metrics")

    add_sheet("Insights", [
        ["#", "Type", "Severity", "Title", "Detail"],
        *[[i+1, ins.get("type",""), ins.get("severity",""), ins.get("title",""), ins.get("detail","")] for i,ins in enumerate(insights)],
    ], title="Automated Insights")

    buf = io.BytesIO()
    wb.save(buf)
    logger.info(f"Excel report generated: {buf.tell():,} bytes")
    return buf.getvalue()